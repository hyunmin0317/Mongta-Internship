import pandas as pd
from bs4 import BeautifulSoup
from urllib.request import urlopen
from openpyxl import load_workbook

load_wb = load_workbook('./link.xlsx',data_only=True)
load_ws = load_wb['Sheet']

urls = []
data = []

for i in range(1,20):
    urls.append(load_ws.cell(i,1).value)

    # columns = []
for url in urls:
    soup = BeautifulSoup(urlopen(url), 'html.parser')
    d = []
    for content in soup.find("div", class_='file-meta-table-pc').find_all("tr"):
        td = (content.find('td').text).replace(u'\xa0', u'').replace(u'\t', u'').replace(u'\n', u'')
        d.append(td)
        # th = (content.find('th').text).replace(u'\xa0', u'').replace(u'\t', u'').replace(u'\n', u'')
        # columns.append(th)
    data.append(d)

df = pd.DataFrame(data, index=urls)
print(df)
df.to_csv('public-data.csv', index=False, encoding='cp949')